from __future__ import annotations

from .models import AbsenceRule
from .models import Entry
from .storage import AbsenceStorage
from .storage import EntryStorage
from collections import defaultdict
from datetime import date
from datetime import datetime
from datetime import time
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict
from typing import Iterable
from typing import List
from typing import Optional
from typing import Sequence
from typing import Tuple

import argparse
import uuid


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import legacy XLSX exports into monthly JSON files.")
    parser.add_argument("xlsx", type=Path, help="Path to the XLSX export")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/entries"),
        help="Directory where JSON month files will be written",
    )
    parser.add_argument("--start-row", type=int, default=6, help="1-based row index where data starts")
    parser.add_argument(
        "--sheet-index",
        type=int,
        default=0,
        help="Index of the sheet to read (0 = first sheet)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing JSON month files instead of merging",
    )
    parser.add_argument(
        "--absences-dir",
        type=Path,
        default=Path("data/absences"),
        help="Directory where detected absences (column W) will be stored as per-year JSON files",
    )
    parser.add_argument(
        "--skip-absences",
        action="store_true",
        help="Ignore column W even if it contains absence information",
    )
    parser.add_argument(
        "--skip-exceptions",
        dest="skip_absences",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser.parse_args(argv)


def normalize_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip().lower() in {"total", "totales"}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        parts = value.split(",", 1)
        if len(parts) == 2:
            value = parts[1].strip()
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Unsupported date value: {value!r}")


def normalize_time(value) -> Optional[time]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%H:%M", "%H.%M"):
            try:
                parsed = datetime.strptime(value, fmt)
                return parsed.time().replace(second=0, microsecond=0)
            except ValueError:
                continue
    if isinstance(value, (int, float)):
        seconds = int(round(float(value) * 24 * 3600))
        hours, remainder = divmod(seconds, 3600)
        minutes = remainder // 60
        return time(hour=hours % 24, minute=minutes % 60)
    raise ValueError(f"Unsupported time value: {value!r}")


def iter_entries_from_row(row, day: date) -> Iterable[Entry]:
    idx = 2  # first column is date, second column ignored
    limit = min(len(row), 16)  # up to column P (index 15)
    while idx < limit:
        start_raw = row[idx]
        end_raw = row[idx + 1] if idx + 1 < len(row) else None
        idx += 2
        start_time = normalize_time(start_raw)
        end_time = normalize_time(end_raw)
        if not start_time or not end_time:
            continue
        start_dt = datetime.combine(day, start_time)
        end_dt = datetime.combine(day, end_time)
        if end_dt <= start_dt:
            continue
        yield Entry(id=str(uuid.uuid4()), start=start_dt, end=end_dt)


def merge_existing(entries_by_month: Dict[str, List[Entry]], output_dir: Path, overwrite: bool) -> None:
    storage = EntryStorage(base_dir=output_dir)
    if overwrite:
        return
    for key, existing in storage.load_all().items():
        entries_by_month.setdefault(key, []).extend(existing)


def write_months(entries_by_month: Dict[str, List[Entry]], output_dir: Path) -> None:
    storage = EntryStorage(base_dir=output_dir)
    for key, entries in entries_by_month.items():
        entries.sort(key=lambda entry: entry.start)
        storage.save_month(key, entries)
        print(f"Wrote {len(entries)} entries to {output_dir / f'{key}.json'}")


def parse_expected_hours(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        hours = float(value)
    except (TypeError, ValueError):
        return None
    return int(round(hours))


def extract_absence(
    row, day: date, expected_value, column_w_value
) -> Optional[Tuple[date, str, Optional[int]]]:
    if column_w_value in (None, ""):
        return None
    reason = str(column_w_value).strip()
    if not reason:
        return None
    hours = parse_expected_hours(expected_value)
    return day, reason, hours


def append_absences(absences: Sequence[Tuple[date, str, Optional[int]]], absences_dir: Path) -> None:
    if not absences:
        return
    storage = AbsenceStorage(base_dir=absences_dir)
    existing_rules = storage.load_all()
    existing_keys = {(rule.start, rule.end or rule.start, rule.reason) for rule in existing_rules}
    added = 0
    for day, reason, hours in absences:
        key = (day, day, reason)
        if key in existing_keys:
            continue
        existing_rules.append(
            AbsenceRule(
                start=day,
                end=day,
                reason=reason,
                hours=hours,
            )
        )
        existing_keys.add(key)
        added += 1
    if added == 0:
        print("No new absences detected.")
        return
    storage.save_rules(existing_rules)
    print(f"Appended {added} absence(s) to {absences_dir}")


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    if not args.xlsx.exists():
        raise SystemExit(f"Input file {args.xlsx} not found")

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    if args.sheet_index >= len(sheet_names):
        raise SystemExit(
            f"Sheet index {args.sheet_index} out of range for workbook with {len(sheet_names)} sheets"
        )
    sheet = wb[sheet_names[args.sheet_index]]

    entries_by_month: Dict[str, List[Entry]] = defaultdict(list)
    absences: List[Tuple[date, str, Optional[int]]] = []
    merge_existing(entries_by_month, args.output_dir, args.overwrite)

    for row in sheet.iter_rows(min_row=args.start_row, values_only=True):
        day = normalize_date(row[0])
        if day is None:
            continue
        for entry in iter_entries_from_row(row, day):
            key = entry.start.strftime("%Y-%m")
            entries_by_month[key].append(entry)
        if not args.skip_absences and len(row) > 22:
            maybe_absence = extract_absence(row, day, row[1], row[22])
            if maybe_absence:
                absences.append(maybe_absence)

    write_months(entries_by_month, args.output_dir)
    if not args.skip_absences:
        append_absences(absences, args.absences_dir)


if __name__ == "__main__":
    main()
